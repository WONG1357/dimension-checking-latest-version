from __future__ import annotations

import re
import unicodedata
from collections import defaultdict
from dataclasses import dataclass
from datetime import date, datetime, time
from typing import Any, Iterable

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


DIMENSION_RE = re.compile(r"^D\d+[A-Za-z]*$")
DATE_RE = re.compile(
    r"(?P<year>20\d{2}|\d{2})[/-](?P<month>\d{1,2})[/-](?P<day>\d{1,2})"
)
TIME_LABEL_RE = re.compile(r"^(?:0|[01]?\d|2[0-3])(?::[0-5]\d){1,2}$")
LABEL_PHRASES = [
    "customerpartno",
    "productpartname",
    "machineno",
    "mcpartno",
    "sono",
    "wono",
    "lotno",
    "date",
    "materialpartnoname",
    "colourpigmentcolorantbatchpartnoname",
    "客户产品编码",
    "产品/配件名称",
    "机台号",
    "mc产品编号",
    "工单号",
    "批号",
    "日期",
    "物料编号名称",
    "原料物料编号名称",
    "色粉种物料编号名称",
]


@dataclass
class DimensionDescriptor:
    display_name: str
    unique_name: str
    header_col: int
    value_col: int
    column_letter: str


@dataclass
class DimensionSpec:
    display_name: str
    unique_name: str
    dimension_column: str
    dimension_column_index: int
    nominal_value: float | None
    positive_tolerance: float | None
    negative_tolerance: float | None
    calculated_upper_limit: float | None
    calculated_lower_limit: float | None
    listed_upper_limit: float | None
    listed_lower_limit: float | None
    upper_limit_match: bool | None
    lower_limit_match: bool | None
    nominal_source_cell: str | None
    positive_tolerance_source_cell: str | None
    negative_tolerance_source_cell: str | None
    listed_upper_source_cell: str | None
    listed_lower_source_cell: str | None


def load_workbook_file(uploaded_file):
    """Load a workbook from a Streamlit upload object or file-like handle."""
    if hasattr(uploaded_file, "seek"):
        uploaded_file.seek(0)
    return load_workbook(uploaded_file, data_only=True, read_only=False)


def normalize_cell_value(value: Any) -> Any:
    """Normalize common Unicode variants and trim whitespace."""
    if value is None:
        return None
    if isinstance(value, str):
        text = unicodedata.normalize("NFKC", value)
        text = text.replace("\u3000", " ").replace("＋", "+").replace("－", "-")
        text = text.replace("\xa0", " ").strip()
        return text or None
    return value


def is_number(value: Any) -> bool:
    if value is None:
        return False
    if isinstance(value, bool):
        return False
    if isinstance(value, (int, float)):
        return True
    text = normalize_cell_value(value)
    if not isinstance(text, str):
        return False
    text = text.replace(",", "")
    try:
        float(text)
        return True
    except (TypeError, ValueError):
        return False


def to_float(value: Any) -> float | None:
    if not is_number(value):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = normalize_cell_value(value)
    if text is None:
        return None
    text = text.replace(",", "")
    try:
        return float(text)
    except (TypeError, ValueError):
        return None


def normalize_search_text(value: Any) -> str:
    value = normalize_cell_value(value)
    if value is None:
        return ""
    text = str(value).lower()
    text = unicodedata.normalize("NFKC", text)
    text = re.sub(r"\s+", "", text)
    text = text.replace("：", ":").replace("(", "").replace(")", "")
    text = text.replace("/", "").replace(".", "").replace("-", "")
    return text


def is_valid_metadata_value(value: Any, field_name: str) -> bool:
    value = normalize_cell_value(value)
    if value is None:
        return False
    if isinstance(value, (datetime, date)):
        return True
    if not isinstance(value, str):
        value = str(value)
    text = value.strip()
    if not text:
        return False
    lowered = normalize_search_text(text)
    if any(phrase in lowered for phrase in LABEL_PHRASES):
        return False
    if field_name == "date":
        return bool(_parse_date_value(value))
    if field_name == "customer_part_no":
        return bool(re.search(r"\d", text)) and not text.lower().startswith(("customer", "客户"))
    if field_name == "product_part_name":
        return not text.lower().startswith(("product", "产品"))
    if field_name == "material_part_no_name":
        return not any(phrase in lowered for phrase in ("materialpartnoname", "colourpigmentcolorantbatchpartnoname", "物料编号名称"))
    if field_name == "machine_no":
        return bool(re.search(r"[a-z]", text, re.I)) and not text.lower().startswith(("machine", "机台"))
    if field_name == "mc_part_no":
        return bool(re.search(r"[a-z0-9]", text, re.I)) and not text.lower().startswith(("mc", "mc产品"))
    if field_name == "so_no":
        return bool(re.search(r"[a-z0-9]", text, re.I))
    if field_name == "wo_no":
        return bool(re.search(r"[a-z0-9]", text, re.I)) and not text.lower().startswith(("wo", "工单"))
    if field_name == "lot_no":
        return bool(re.search(r"\d", text)) and not text.lower().startswith(("lot", "批号"))
    return True


def resolve_cell_value(ws, row: int, col: int, merged_map: dict[tuple[int, int], Any]) -> Any:
    value = ws.cell(row=row, column=col).value
    if value is not None:
        return normalize_cell_value(value)
    return merged_map.get((row, col))


def build_merged_map(ws) -> dict[tuple[int, int], Any]:
    merged_map: dict[tuple[int, int], Any] = {}
    for merged_range in ws.merged_cells.ranges:
        min_row, min_col, max_row, max_col = (
            merged_range.min_row,
            merged_range.min_col,
            merged_range.max_row,
            merged_range.max_col,
        )
        top_left = normalize_cell_value(ws.cell(min_row, min_col).value)
        for r in range(min_row, max_row + 1):
            for c in range(min_col, max_col + 1):
                merged_map[(r, c)] = top_left
    return merged_map


def _parse_time_value(value: Any) -> str | None:
    value = normalize_cell_value(value)
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.strftime("%H:%M")
    if isinstance(value, time):
        return value.strftime("%H:%M")
    if isinstance(value, (int, float)):
        if float(value) == 0:
            return "00:00"
        if 0 <= float(value) < 1:
            total_minutes = round(float(value) * 24 * 60)
            return f"{total_minutes // 60:02d}:{total_minutes % 60:02d}"
    if isinstance(value, str):
        text = value.strip()
        if TIME_LABEL_RE.match(text):
            parts = text.split(":")
            hour = int(parts[0])
            minute = int(parts[1]) if len(parts) > 1 else 0
            return f"{hour:02d}:{minute:02d}"
    return None


def _parse_date_value(value: Any) -> str | None:
    value = normalize_cell_value(value)
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, date):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, str):
        text = value.strip()
        match = DATE_RE.search(text)
        if match:
            year = match.group("year")
            if len(year) == 2:
                year = f"20{year}"
            month = int(match.group("month"))
            day = int(match.group("day"))
            return f"{int(year):04d}-{month:02d}-{day:02d}"
    return None


def _row_contains_any(values: list[Any], needles: Iterable[str]) -> bool:
    text = " ".join(normalize_search_text(v) for v in values if v is not None)
    return any(needle in text for needle in needles)


def _cell_contains_any(value: Any, needles: Iterable[str]) -> bool:
    text = normalize_search_text(value)
    return any(needle in text for needle in needles)


def _find_first_meaningful_rightward(
    ws,
    row: int,
    col: int,
    merged_map: dict[tuple[int, int], Any],
    max_cols: int = 10,
    field_name: str | None = None,
) -> Any:
    for offset in range(1, max_cols + 1):
        value = resolve_cell_value(ws, row, col + offset, merged_map)
        if value is None:
            continue
        if field_name is not None and not is_valid_metadata_value(value, field_name):
            continue
        return value
    return None


def _search_metadata_candidate(
    ws,
    label_row: int,
    label_col: int,
    merged_map: dict[tuple[int, int], Any],
    field_name: str,
) -> Any:
    search_rows = [label_row, label_row + 1, label_row + 2, label_row + 3]
    for row_idx in search_rows:
        if row_idx < 1 or row_idx > ws.max_row:
            continue
        for col_idx in range(label_col + 1, min(ws.max_column, label_col + 12) + 1):
            value = resolve_cell_value(ws, row_idx, col_idx, merged_map)
            if is_valid_metadata_value(value, field_name):
                return value
    return None


def _first_col_after_box(ws, row: int, col: int) -> int:
    for merged_range in ws.merged_cells.ranges:
        if merged_range.min_row <= row <= merged_range.max_row and merged_range.min_col <= col <= merged_range.max_col:
            return merged_range.max_col + 1
    return col + 1


def _parse_date_from_context(ws, label_row: int, label_col: int, merged_map: dict[tuple[int, int], Any]) -> str | None:
    numbers: list[int] = []
    for row_idx in [label_row, label_row + 1, label_row + 2, label_row + 3]:
        if row_idx < 1 or row_idx > ws.max_row:
            continue
        for col_idx in range(label_col, min(ws.max_column, label_col + 12) + 1):
            value = resolve_cell_value(ws, row_idx, col_idx, merged_map)
            if isinstance(value, datetime):
                return value.strftime("%Y-%m-%d")
            if isinstance(value, date):
                return value.strftime("%Y-%m-%d")
            if isinstance(value, (int, float)) and float(value).is_integer():
                numbers.append(int(value))
            elif isinstance(value, str):
                parsed = _parse_date_value(value)
                if parsed:
                    return parsed
    if len(numbers) >= 3:
        year, month, day = numbers[:3]
        if year < 100:
            year += 2000
        if 1900 <= year <= 2100 and 1 <= month <= 12 and 1 <= day <= 31:
            return f"{year:04d}-{month:02d}-{day:02d}"
    return None


def _extract_metadata_from_label_block(
    ws,
    label_row: int,
    label_col: int,
    merged_map: dict[tuple[int, int], Any],
    field_name: str,
) -> Any:
    start_col = _first_col_after_box(ws, label_row, label_col)
    candidates: list[Any] = []
    for row_idx in [label_row, label_row + 1, label_row + 2, label_row + 3]:
        if row_idx < 1 or row_idx > ws.max_row:
            continue
        for col_idx in range(start_col, min(ws.max_column, start_col + 12) + 1):
            value = resolve_cell_value(ws, row_idx, col_idx, merged_map)
            if is_valid_metadata_value(value, field_name):
                candidates.append(value)
    if candidates:
        return candidates[0]
    return None


def extract_header_info(ws) -> dict[str, Any]:
    """Extract header metadata with keyword search rather than fixed cells."""
    merged_map = build_merged_map(ws)
    headers = {
        "date": None,
        "customer_part_no": None,
        "product_part_name": None,
        "material_part_no_name": None,
        "machine_no": None,
        "mc_part_no": None,
        "so_no": None,
        "wo_no": None,
        "lot_no": None,
    }
    labels = {
        "customer_part_no": ["customerpartno", "客户产品编码"],
        "product_part_name": ["productpartname", "产品/配件名称"],
        "material_part_no_name": ["materialpartnoname", "colourpigmentcolorantbatchpartnoname"],
        "machine_no": ["machineno", "机台号"],
        "mc_part_no": ["mcpartno", "mc产品编号"],
        "so_no": ["sono", "销售单号"],
        "wo_no": ["wono", "工单号"],
        "lot_no": ["lotno", "批号"],
    }

    for row in ws.iter_rows():
        for cell in row:
            value = resolve_cell_value(ws, cell.row, cell.column, merged_map)
            if value is None:
                continue
            text = normalize_search_text(value)
            if not text:
                continue
            if headers["date"] is None and ("date" in text or "日期" in str(value)):
                parsed_date = _parse_date_from_context(ws, cell.row, cell.column, merged_map)
                if parsed_date:
                    headers["date"] = parsed_date
            for key, needle_list in labels.items():
                if headers[key] is not None:
                    continue
                if any(needle in text for needle in needle_list):
                    candidate = _extract_metadata_from_label_block(ws, cell.row, cell.column, merged_map, key)
                    if candidate is not None:
                        headers[key] = normalize_cell_value(candidate)

    if headers["date"] is None:
        for row in ws.iter_rows():
            for cell in row:
                candidate = resolve_cell_value(ws, cell.row, cell.column, merged_map)
                parsed_date = _parse_date_value(candidate)
                if parsed_date:
                    headers["date"] = parsed_date
                    break
            if headers["date"]:
                break

    return headers


def find_dimension_header_row(ws) -> tuple[int | None, list[DimensionDescriptor]]:
    """Locate the row containing the dimension names and return unique descriptors."""
    best_row = None
    best_hits: list[tuple[int, int, str]] = []
    for row in ws.iter_rows():
        hits: list[tuple[int, int, str]] = []
        for cell in row:
            value = normalize_cell_value(cell.value)
            if isinstance(value, str) and DIMENSION_RE.match(value.strip()):
                hits.append((cell.row, cell.column, value.strip()))
        if len(hits) > len(best_hits):
            best_hits = hits
            best_row = row[0].row
    if not best_hits:
        return None, []

    counts: defaultdict[str, int] = defaultdict(int)
    descriptors: list[DimensionDescriptor] = []
    for _, col_idx, display_name in best_hits:
        counts[display_name] += 1
        unique_name = display_name if counts[display_name] == 1 else f"{display_name}_{counts[display_name]}"
        value_col = col_idx + 1
        descriptors.append(
            DimensionDescriptor(
                display_name=display_name,
                unique_name=unique_name,
                header_col=col_idx,
                value_col=value_col,
                column_letter=get_column_letter(value_col),
            )
        )
    descriptors.sort(key=lambda item: item.header_col)
    return best_row, descriptors


def find_equipment_row(ws) -> int | None:
    merged_map = build_merged_map(ws)
    keywords = ["equipmentno", "测量工具编号"]
    for row in ws.iter_rows():
        row_values = [resolve_cell_value(ws, c.row, c.column, merged_map) for c in row]
        if _row_contains_any(row_values, keywords):
            return row[0].row
    return None


def _nearby_sign_exists(ws, row: int, col: int, merged_map: dict[tuple[int, int], Any], sign: str) -> bool:
    for c in range(max(1, col - 2), min(ws.max_column, col + 1) + 1):
        value = resolve_cell_value(ws, row, c, merged_map)
        if normalize_cell_value(value) == sign:
            return True
    return False


def _find_value_in_row(
    ws,
    row: int,
    value_col: int,
    merged_map: dict[tuple[int, int], Any],
) -> tuple[float | None, str | None]:
    value = resolve_cell_value(ws, row, value_col, merged_map)
    if is_number(value):
        return to_float(value), f"{get_column_letter(value_col)}{row}"
    return None, None


def extract_dimension_specs(
    ws,
    dimension_header_row: int,
    dimension_columns: list[DimensionDescriptor],
) -> dict[str, DimensionSpec]:
    """Extract nominal, tolerances, and listed limits for every detected dimension."""
    merged_map = build_merged_map(ws)
    equipment_row = find_equipment_row(ws) or ws.max_row
    specs: dict[str, DimensionSpec] = {}

    for descriptor in dimension_columns:
        nominal_value = positive_tolerance = negative_tolerance = None
        listed_upper = listed_lower = None
        nominal_source = positive_source = negative_source = upper_source = lower_source = None

        candidate_rows: list[tuple[int, float]] = []
        for row_idx in range(dimension_header_row + 1, equipment_row):
            value, cell_ref = _find_value_in_row(ws, row_idx, descriptor.value_col, merged_map)
            if value is not None:
                candidate_rows.append((row_idx, value))
                if nominal_value is None:
                    nominal_value = value
                    nominal_source = cell_ref

        if nominal_value is None:
            specs[descriptor.unique_name] = DimensionSpec(
                display_name=descriptor.display_name,
                unique_name=descriptor.unique_name,
                dimension_column=descriptor.column_letter,
                dimension_column_index=descriptor.value_col,
                nominal_value=None,
                positive_tolerance=None,
                negative_tolerance=None,
                calculated_upper_limit=None,
                calculated_lower_limit=None,
                listed_upper_limit=None,
                listed_lower_limit=None,
                upper_limit_match=None,
                lower_limit_match=None,
                nominal_source_cell=None,
                positive_tolerance_source_cell=None,
                negative_tolerance_source_cell=None,
                listed_upper_source_cell=None,
                listed_lower_source_cell=None,
            )
            continue

        for row_idx in range(dimension_header_row + 1, equipment_row):
            value, cell_ref = _find_value_in_row(ws, row_idx, descriptor.value_col, merged_map)
            if value is None:
                continue
            if positive_tolerance is None and _nearby_sign_exists(ws, row_idx, descriptor.value_col, merged_map, "+"):
                positive_tolerance = value
                positive_source = cell_ref
                continue
            if negative_tolerance is None and _nearby_sign_exists(ws, row_idx, descriptor.value_col, merged_map, "-"):
                negative_tolerance = value
                negative_source = cell_ref
                continue

        numeric_rows_after_tol = [
            (row_idx, value)
            for row_idx, value in candidate_rows
            if row_idx > dimension_header_row + 1
        ]
        if positive_tolerance is not None and negative_tolerance is not None:
            tolerance_end_row = max(
                [
                    row_idx
                    for row_idx in range(dimension_header_row + 1, equipment_row)
                    if _nearby_sign_exists(ws, row_idx, descriptor.value_col, merged_map, "+")
                    or _nearby_sign_exists(ws, row_idx, descriptor.value_col, merged_map, "-")
                ],
                default=dimension_header_row,
            )
            limit_candidates = [
                (row_idx, value)
                for row_idx, value in numeric_rows_after_tol
                if row_idx > tolerance_end_row
            ]
            if limit_candidates:
                listed_upper = limit_candidates[0][1]
                upper_source = f"{descriptor.column_letter}{limit_candidates[0][0]}"
            if len(limit_candidates) > 1:
                listed_lower = limit_candidates[1][1]
                lower_source = f"{descriptor.column_letter}{limit_candidates[1][0]}"

        calculated_upper = nominal_value + positive_tolerance if positive_tolerance is not None else None
        calculated_lower = nominal_value - negative_tolerance if negative_tolerance is not None else None
        upper_match = (
            abs(calculated_upper - listed_upper) <= 0.001
            if calculated_upper is not None and listed_upper is not None
            else None
        )
        lower_match = (
            abs(calculated_lower - listed_lower) <= 0.001
            if calculated_lower is not None and listed_lower is not None
            else None
        )

        specs[descriptor.unique_name] = DimensionSpec(
            display_name=descriptor.display_name,
            unique_name=descriptor.unique_name,
            dimension_column=descriptor.column_letter,
            dimension_column_index=descriptor.value_col,
            nominal_value=nominal_value,
            positive_tolerance=positive_tolerance,
            negative_tolerance=negative_tolerance,
            calculated_upper_limit=calculated_upper,
            calculated_lower_limit=calculated_lower,
            listed_upper_limit=listed_upper,
            listed_lower_limit=listed_lower,
            upper_limit_match=upper_match,
            lower_limit_match=lower_match,
            nominal_source_cell=nominal_source,
            positive_tolerance_source_cell=positive_source,
            negative_tolerance_source_cell=negative_source,
            listed_upper_source_cell=upper_source,
            listed_lower_source_cell=lower_source,
        )

    return specs


def _is_blank_row(values: list[Any]) -> bool:
    return all(v is None or str(v).strip() == "" for v in values)


def _find_inspection_time_in_row(values: list[Any]) -> str | None:
    for value in values[:4]:
        parsed = _parse_time_value(value)
        if parsed is not None:
            return parsed
    return None


def extract_sample_rows(
    ws,
    equipment_row: int,
    dimension_specs: dict[str, DimensionSpec],
) -> tuple[list[dict[str, Any]], int]:
    """Extract sample measurements below the equipment row."""
    merged_map = build_merged_map(ws)
    records: list[dict[str, Any]] = []
    inspection_rows = 0
    blank_streak = 0
    started = False
    remarks_keywords = ["备注", "remarks"]

    for row_idx in range(equipment_row + 1, ws.max_row + 1):
        row_values = [resolve_cell_value(ws, row_idx, col, merged_map) for col in range(1, ws.max_column + 1)]
        row_text = " ".join(normalize_search_text(v) for v in row_values if v is not None)
        if any(keyword in row_text for keyword in remarks_keywords):
            break
        if _is_blank_row(row_values):
            blank_streak += 1
            if started and blank_streak >= 2:
                break
            continue
        blank_streak = 0

        inspection_time = _find_inspection_time_in_row(row_values)
        if inspection_time is None:
            continue
        started = True
        inspection_rows += 1

        inspection_quantity = None
        defective_quantity = None
        if len(row_values) >= 2 and is_number(row_values[1]):
            inspection_quantity = to_float(row_values[1])
        if len(row_values) >= 3 and is_number(row_values[2]):
            defective_quantity = to_float(row_values[2])

        for spec in dimension_specs.values():
            measurement = resolve_cell_value(ws, row_idx, spec.dimension_column_index, merged_map)
            if not is_number(measurement):
                continue
            measurement_value = to_float(measurement)
            if measurement_value is None:
                continue
            status = validate_measurement(
                measurement_value,
                spec.calculated_lower_limit,
                spec.calculated_upper_limit,
            )
            records.append(
                {
                    "inspection_row": row_idx,
                    "inspection_time": inspection_time,
                    "inspection_quantity": inspection_quantity,
                    "defective_quantity": defective_quantity,
                    "measurement_value": measurement_value,
                    "source_cell": f"{spec.dimension_column}{row_idx}",
                    "status": status,
                    "spec_unique_name": spec.unique_name,
                }
            )

    return records, inspection_rows


def validate_measurement(
    measurement: float | None,
    lower_limit: float | None,
    upper_limit: float | None,
) -> str:
    if measurement is None or lower_limit is None or upper_limit is None:
        return "SPEC_MISSING"
    return "PASS" if lower_limit <= measurement <= upper_limit else "FAIL"


def parse_workbook(uploaded_file) -> dict[str, Any]:
    workbook = load_workbook_file(uploaded_file)
    all_records: list[dict[str, Any]] = []
    sheet_summaries: list[dict[str, Any]] = []
    dimension_rows: list[dict[str, Any]] = []
    warnings: list[str] = []
    errors: list[dict[str, Any]] = []

    for ws in workbook.worksheets:
        try:
            header_info = extract_header_info(ws)
            dimension_header_row, descriptors = find_dimension_header_row(ws)
            if dimension_header_row is None or not descriptors:
                warnings.append(f"{ws.title}: no dimension header row detected.")
                continue
            equipment_row = find_equipment_row(ws)
            if equipment_row is None:
                warnings.append(f"{ws.title}: equipment row not detected; using sheet end.")
                equipment_row = ws.max_row
            specs = extract_dimension_specs(ws, dimension_header_row, descriptors)
            sample_records, inspection_rows = extract_sample_rows(ws, equipment_row, specs)

            detail_rows = 0
            fail_rows = 0
            pass_rows = 0
            for record in sample_records:
                spec = specs.get(record["spec_unique_name"])
                if spec is None:
                    continue
                difference_from_nominal = (
                    record["measurement_value"] - spec.nominal_value
                    if spec.nominal_value is not None
                    else None
                )
                row = {
                    "sheet_name": ws.title,
                    "date": header_info.get("date"),
                    "customer_part_no": header_info.get("customer_part_no"),
                    "product_part_name": header_info.get("product_part_name"),
                    "material_part_no_name": header_info.get("material_part_no_name"),
                    "machine_no": header_info.get("machine_no"),
                    "mc_part_no": header_info.get("mc_part_no"),
                    "so_no": header_info.get("so_no"),
                    "wo_no": header_info.get("wo_no"),
                    "lot_no": header_info.get("lot_no"),
                    "inspection_time": record["inspection_time"],
                    "inspection_quantity": record["inspection_quantity"],
                    "defective_quantity": record["defective_quantity"],
                    "dimension_display_name": spec.display_name,
                    "dimension_unique_name": spec.unique_name,
                    "dimension_column": spec.dimension_column,
                    "nominal_value": spec.nominal_value,
                    "positive_tolerance": spec.positive_tolerance,
                    "negative_tolerance": spec.negative_tolerance,
                    "calculated_upper_limit": spec.calculated_upper_limit,
                    "calculated_lower_limit": spec.calculated_lower_limit,
                    "listed_upper_limit": spec.listed_upper_limit,
                    "listed_lower_limit": spec.listed_lower_limit,
                    "upper_limit_match": spec.upper_limit_match,
                    "lower_limit_match": spec.lower_limit_match,
                    "measurement_value": record["measurement_value"],
                    "difference_from_nominal": difference_from_nominal,
                    "status": record["status"],
                    "source_cell": record["source_cell"],
                }
                all_records.append(row)
                detail_rows += 1
                if row["status"] == "PASS":
                    pass_rows += 1
                elif row["status"] == "FAIL":
                    fail_rows += 1

            for spec in specs.values():
                dimension_rows.append(
                    {
                        "sheet_name": ws.title,
                        "dimension_display_name": spec.display_name,
                        "dimension_unique_name": spec.unique_name,
                        "dimension_column": spec.dimension_column,
                        "nominal_value": spec.nominal_value,
                        "positive_tolerance": spec.positive_tolerance,
                        "negative_tolerance": spec.negative_tolerance,
                        "calculated_upper_limit": spec.calculated_upper_limit,
                        "calculated_lower_limit": spec.calculated_lower_limit,
                        "listed_upper_limit": spec.listed_upper_limit,
                        "listed_lower_limit": spec.listed_lower_limit,
                        "upper_limit_match": spec.upper_limit_match,
                        "lower_limit_match": spec.lower_limit_match,
                        "nominal_source_cell": spec.nominal_source_cell,
                        "positive_tolerance_source_cell": spec.positive_tolerance_source_cell,
                        "negative_tolerance_source_cell": spec.negative_tolerance_source_cell,
                        "listed_upper_source_cell": spec.listed_upper_source_cell,
                        "listed_lower_source_cell": spec.listed_lower_source_cell,
                    }
                )

            sheet_summaries.append(
                {
                    "sheet_name": ws.title,
                    "date": header_info.get("date"),
                    "customer_part_no": header_info.get("customer_part_no"),
                    "product_part_name": header_info.get("product_part_name"),
                    "machine_no": header_info.get("machine_no"),
                    "mc_part_no": header_info.get("mc_part_no"),
                    "wo_no": header_info.get("wo_no"),
                    "lot_no": header_info.get("lot_no"),
                    "inspection_record_count": inspection_rows,
                    "measurement_count": detail_rows,
                    "pass_count": pass_rows,
                    "fail_count": fail_rows,
                    "overall_result": "FAIL" if fail_rows else "PASS" if detail_rows else "NO DATA",
                    "dimension_count": len(specs),
                }
            )
        except Exception as exc:  # pragma: no cover - sheet-level resilience
            warnings.append(f"{ws.title}: parsing failed - {exc}")
            errors.append({"sheet_name": ws.title, "error": str(exc)})

    detail_df_columns = [
        "sheet_name",
        "date",
        "customer_part_no",
        "product_part_name",
        "material_part_no_name",
        "machine_no",
        "mc_part_no",
        "so_no",
        "wo_no",
        "lot_no",
        "inspection_time",
        "inspection_quantity",
        "defective_quantity",
        "dimension_display_name",
        "dimension_unique_name",
        "dimension_column",
        "nominal_value",
        "positive_tolerance",
        "negative_tolerance",
        "calculated_upper_limit",
        "calculated_lower_limit",
        "listed_upper_limit",
        "listed_lower_limit",
        "upper_limit_match",
        "lower_limit_match",
        "measurement_value",
        "difference_from_nominal",
        "status",
        "source_cell",
    ]

    return {
        "workbook": workbook,
        "detail_rows": all_records,
        "detail_df_columns": detail_df_columns,
        "sheet_summaries": sheet_summaries,
        "dimension_summaries": dimension_rows,
        "warnings": warnings,
        "errors": errors,
    }
