from __future__ import annotations

from io import BytesIO
from typing import Any

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)
PASS_FILL = PatternFill("solid", fgColor="E2F0D9")
FAIL_FILL = PatternFill("solid", fgColor="F4CCCC")
SPEC_MISSING_FILL = PatternFill("solid", fgColor="FFF2CC")


def _auto_widths(ws, dataframe: pd.DataFrame, max_width: int = 36) -> None:
    for idx, col in enumerate(dataframe.columns, start=1):
        values = [str(col)] + ["" if pd.isna(v) else str(v) for v in dataframe[col].head(200).tolist()]
        width = min(max(len(v) for v in values) + 2, max_width)
        ws.column_dimensions[get_column_letter(idx)].width = max(10, width)


def _write_dataframe(ws, dataframe: pd.DataFrame, highlight_status: bool = False) -> None:
    for col_idx, col_name in enumerate(dataframe.columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    for row_idx, row in enumerate(dataframe.itertuples(index=False), start=2):
        for col_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=None if pd.isna(value) else value)
            if highlight_status and "status" in dataframe.columns:
                status = dataframe.iloc[row_idx - 2]["status"]
                if status == "PASS":
                    cell.fill = PASS_FILL
                elif status == "FAIL":
                    cell.fill = FAIL_FILL
                elif status == "SPEC_MISSING":
                    cell.fill = SPEC_MISSING_FILL
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    _auto_widths(ws, dataframe)


def export_validation_report(
    detail_df: pd.DataFrame,
    sheet_summary_df: pd.DataFrame,
    overall_summary: dict[str, Any],
    dimension_summary_df: pd.DataFrame,
) -> bytes:
    wb = Workbook()
    default_ws = wb.active
    wb.remove(default_ws)

    overview = wb.create_sheet("Overall Summary")
    overview_rows = [
        ["Metric", "Value"],
        ["Total Sheets", overall_summary.get("total_sheets", 0)],
        ["Total Records", overall_summary.get("total_records", 0)],
        ["Total Measurements", overall_summary.get("total_measurements", 0)],
        ["PASS Count", overall_summary.get("pass_count", 0)],
        ["FAIL Count", overall_summary.get("fail_count", 0)],
        ["SPEC_MISSING Count", overall_summary.get("spec_missing_count", 0)],
        ["Overall Result", overall_summary.get("overall_result", "NO DATA")],
    ]
    for r_idx, row in enumerate(overview_rows, start=1):
        for c_idx, value in enumerate(row, start=1):
            cell = overview.cell(row=r_idx, column=c_idx, value=value)
            if r_idx == 1:
                cell.fill = HEADER_FILL
                cell.font = HEADER_FONT
    overview.column_dimensions["A"].width = 22
    overview.column_dimensions["B"].width = 18

    detail_ws = wb.create_sheet("Validation Detail")
    _write_dataframe(detail_ws, detail_df, highlight_status=True)

    sheet_ws = wb.create_sheet("Sheet Summary")
    _write_dataframe(sheet_ws, sheet_summary_df, highlight_status=False)

    dimension_ws = wb.create_sheet("Dimension Summary")
    _write_dataframe(dimension_ws, dimension_summary_df, highlight_status=False)

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio.getvalue()

